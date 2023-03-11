"""
This is a program that gets questions from Open Trivia Database and prints them to an xlsx file.
"""

import datetime
from html import unescape
import random
import shutil
import openpyxl
import requests


def main():
    """This is the main function."""
    date_time: str = f"{datetime.date.today()}-{datetime.datetime.now().time().strftime('%H-%M')}"
    new_file_name: str = f"{date_time}-kahoot-quiz.xlsx"
    create_new_xlsx_file(new_file_name)

    request: object = get_questions_from_db(20)
    question_list: QuestionList = process_json_for_questions(request)

    print_to_xlsx(question_list, new_file_name)


def create_new_xlsx_file(new_file_name: str):
    '''This function will create a new xlsx file from template'''
    shutil.copy2('Kahoot-Quiz-Spreadsheet-Template.xlsx', f'generated_quizzes/{new_file_name}')


def print_to_xlsx(question_list: object, new_file_name: str):
    '''This function will print the questions to the xlsx file'''
    file: object = openpyxl.load_workbook(f'generated_quizzes/{new_file_name}')
    sheet: object = file.active

    num_of_questions: int = question_list.num_of_questions()
    for i in range(9, num_of_questions + 9):
        sheet.cell(row=i, column=2).value = question_list.question_list[i-9].question
        sheet.cell(row=i, column=3).value = question_list.question_list[i-9].answers[0]
        sheet.cell(row=i, column=4).value = question_list.question_list[i-9].answers[1]
        sheet.cell(row=i, column=5).value = question_list.question_list[i-9].answers[2]
        sheet.cell(row=i, column=6).value = question_list.question_list[i-9].answers[3]
        sheet.cell(row=i, column=7).value = "60"
        sheet.cell(row=i, column=8).value = question_list.question_list[i-9].correct_answer_number

    file.save(f'generated_quizzes/{new_file_name}')


def process_json_for_questions(request : object) -> object:
    '''This function will process the json data and create a list of questions'''
    question_list: QuestionList = QuestionList()

    for result in request.json()['results']:
        question_answers: str = unescape(result['incorrect_answers'])
        answer_position: int = random.randint(0, 3)
        question_answers.insert(answer_position, unescape(result['correct_answer']))

        full_question: Question = Question(
            unescape(result['question']),
            question_answers, answer_position+1
            )
        question_list.add_question(full_question)

    return question_list


def get_questions_from_db(num_of_questions : int):
    '''This function will get the questions from the database'''
    headers: dict = {'Accept': 'application/json'}
    request: object = requests.get(
        f'https://opentdb.com/api.php?amount={num_of_questions}&category=15&type=multiple',
        headers=headers,
        timeout=20
    )

    return request


class QuestionList:
    """Creates a list of questions."""

    def __init__(self):
        self.question_list = []

    def __str__(self) -> str:
        return f"{self.question_list}"

    def add_question(self, question):
        '''This function will add a question to the list of questions'''
        self.question_list.append(question)

    def print_questions(self):
        '''This function will print the questions'''
        for question in self.question_list:
            print(question)

    def num_of_questions(self) -> int:
        '''This function will return the number of questions in the list'''
        return len(self.question_list)

    def __iter__(self) -> object:
        return QuestionListIter(self)


class QuestionListIter:
    """Creates an iterator for the question list."""

    def __init__(self, question_list):
        self._questions = question_list.question_list
        self._num_of_questions = question_list.num_of_questions()
        self._current_index = 0

    def __iter__(self) -> object:
        return self

    def __next__(self) -> object:
        if self._current_index < self._num_of_questions:
            self._current_index += 1
            return self._questions[self._current_index - 1]

        raise StopIteration


class Question:
    """Creates a question object."""
    def __init__(self, question, answers, correct_answer_number):
        self.question = question
        self.answers = answers
        self.correct_answer_number = correct_answer_number

    def __str__(self) -> str:
        return f"Q:{self.question} A: {self.answers} Correct Answer: {self.correct_answer_number}"

    def __repr__(self) -> str:
        return f"Q:{self.question} A: {self.answers} Correct Answer: {self.correct_answer_number}"




if __name__ == "__main__":
    main()
